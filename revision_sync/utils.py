from datetime import datetime
from django.db.models import F
from io import BytesIO
from openpyxl import load_workbook
import logging
import urllib.request

from revision_app import models

logger = logging.getLogger(__name__)


# Ref : https://stackoverflow.com/questions/20635778/using-openpyxl-to-read-file-from-memory
def load_workbook_from_url(url):
    file = urllib.request.urlopen(url).read()
    return load_workbook(filename = BytesIO(file))

def get_sheet_from_workbook(book,sheet_name):
    return book[sheet_name]

def parse_sheet_for_revisions(sheet):
    mindmaps_list = []
    for row in sheet.iter_rows(2):
        if row[2].value:
            revision_dates = [r.value for r in row[5:14]]
            mindmap = {
                "topic": row[2].value,
                "category": row[1].value,
                "created": row[14].value,
                "revisionLevel": row[3].value ,
                "nextLevelStartDate": row[4].value,
                "revisionDates": revision_dates,
            }
            mindmaps_list.append(mindmap)
    return mindmaps_list

def __create_categories(categories):
    logger.info(f"Creating categories : {str(categories)}")
    models.Category.objects.bulk_create([models.Category(title=title) for title in categories])

"""
Problem : Creating mindmap from mindmaps_list
mindmap item has : topic, category, created, revisionLevel, nextLevelStartDate, revisionDates

Nest mindmaps as
    categories > mindmaps > revisionDates

Iterate over category, for each category check if category exists, if not then
create category. 
After that iterate over mindmaps in this category, and check mindmap exists for this mindmap in this category,
if not then create mindmap and revision group.
After that iterate over revisionDates for this mindmap and check if revision Item exists for every date,
if not then create revision item for that revision group.
"""

def __create_mindmaps(mindmaps_list) :
    categories_stored = models.Category.objects.all().values()
    revLevels_stored = models.RevisionLevel.objects.all().values()

    categories = {}
    for mindmap in mindmaps_list:
        ctg = mindmap['category']
        if ctg in categories:
            categories[ctg].append(mindmap)
        else: categories[ctg] = [mindmap]

    def get_category_for_mindmap(m):
        for ctg in categories_stored:
            if m['category'] == ctg['title']:
                return models.Category(id=ctg['id'], title=ctg['title'])

    def get_revision_level(m):        
        lvl = int(m['revisionLevel'])
        lvl1 = [item for item in revLevels_stored if item['level'] == f"Level {lvl}"][0]
        return models.RevisionLevel(id=lvl1['id'], level=lvl1['level'])
    
    def get_date(date):
        if date:
            return datetime.strptime(date, "%d.%m.%Y")
        return datetime.now()

    for ctg in categories:
        for mMap in categories[ctg]:
            # topic, category, created, revisionLevel, nextLevelStartDate, revisionDates
            m = models.Mindmap(title=mMap['topic'], description=mMap['topic'], 
                category=get_category_for_mindmap(mMap), image_link="https://to.do", 
                creation_date=get_date(mMap['created']))
            m.save()

            if mMap['revisionLevel']:
                group = models.RevisionGroup(mindmap=m, revision_level=get_revision_level(mMap))
                group.save()
                for date in mMap['revisionDates']:
                    if date:
                        if len(date.strip()) == 10:
                            date_obj = datetime.strptime(date, "%d.%m.%Y")
                            models.RevisionItem(revision_group=group, date=date_obj, revision_done=True).save()
                        else:
                            date_obj = datetime.strptime(date, "T %d.%m.%Y")
                            models.RevisionItem(revision_group=group, date=date_obj, revision_done=False).save()


def create_mindmaps_from_list(mindmaps_list) :
    # Cleaning list
    mindmaps_list = [item for item in mindmaps_list if item['topic'] is not None]

    categories_list = [item['category'] for item in mindmaps_list]
    categories_stored_qs = models.Category.objects.filter(title__in=categories_list).values()
    categories_stored = [item['title'] for item in categories_stored_qs ]
    
    # Checking for category
    categories_to_create = []
    for mindmap in mindmaps_list:
        if mindmap['category'] not in categories_stored and\
            mindmap['category'] not in categories_to_create:
            categories_to_create.append(mindmap['category'])
    if len(categories_to_create) > 0 : __create_categories(categories_to_create)

    # Checking for mindmaps
    titles_list = [item['topic'] for item in mindmaps_list]
    mindmaps_stored = models.Mindmap.objects.select_related('category')\
            .filter(title__in=titles_list).annotate(category_name=F('category__title')).values()
    # Grouping for easy iteration over category and mindmaps
    mindmaps_category_grouped = {}
    for mindmap in mindmaps_stored:
        if mindmap['category_name'] in mindmaps_category_grouped:
            mindmaps_category_grouped[mindmap['category_name']].append(mindmap['title'])
        else: mindmaps_category_grouped[mindmap['category_name']] = [mindmap['title']]

    mindmaps_to_create = []
    for mindmap in mindmaps_list:
        if mindmap['category'] in mindmaps_category_grouped:
            if mindmap['topic'] not in mindmaps_category_grouped[mindmap['category']]:
                mindmaps_to_create.append(mindmap)
        else: mindmaps_to_create.append(mindmap)

    if len(mindmaps_to_create) > 0 : __create_mindmaps(mindmaps_to_create)

if __name__ == "__main__":
    print("Testing utils.")


    # link = "https://docs.google.com/spreadsheets/d/<spreadsheet-id>/export?format=xlsx"

    # book = load_workbook_from_url(link)

    # sheet = book['Sheet1']

    # mindmaps_list = parse_sheet_for_revisions(sheet)
    # print(mindmaps_list[-1])
    # print(len(mindmaps_list))
